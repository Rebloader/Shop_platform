import openpyxl
from datetime import datetime
from pathlib import Path

from src.config import BASE_DIR


async def create_excel_file_with_order_info(order_data: dict) -> str:
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet['A1'] = 'Заказчик'
    sheet['B1'] = 'Дата заказа'
    sheet['C1'] = 'Товар'
    sheet['D1'] = 'Кол-во'
    sheet['E1'] = 'Цена за единицу'
    sheet['F1'] = 'Итого'

    row = 2
    sheet.merge_cells('A2:A' + str(len(order_data['items']) + 1))
    sheet['A2'] = order_data['dealer_name']
    sheet.merge_cells('B2:B' + str(len(order_data['items']) + 1))
    sheet['B2'] = order_data['created_at']
    for item in order_data['items']:
        sheet['C' + str(row)] = item['product_name']
        sheet['D' + str(row)] = item['quantity']
        sheet['E' + str(row)] = item['price']
        row += 1
    sheet['F' + str(row)] = order_data['total_price']

    file_name = f"order_{order_data['dealer_name']}_{datetime.now().strftime('%Y-%m-%d-%H%M%S')}.xlsx"
    file_path = f'{BASE_DIR}/documents/{file_name}'
    workbook.save(file_path)
    return file_name

